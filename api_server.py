from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, Field
import uvicorn
import os
import tempfile
import asyncio
import ast
from fractions import Fraction
import hashlib
import logging
import math
import time
import uuid
import re
from contextlib import asynccontextmanager
from typing import Optional, Dict, List, Any
import json
from statistics import mean
from types import SimpleNamespace
from urllib import error as urllib_error
from urllib import request as urllib_request

try:
    from dotenv import load_dotenv
    load_dotenv()
except Exception:
    pass
# --- Logging ---
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
)
logger = logging.getLogger("MathMendAPI")

# ---------------------------------------------------------------
# CONFIG — qwen3.5:397b-cloud via Ollama
# Run first: ollama run qwen3.5:397b-cloud
# ---------------------------------------------------------------
MODEL_NAME = os.getenv("OLLAMA_MODEL", "qwen2.5:7b")


class _OllamaChatCompletions:
    def __init__(self, base_url: str, timeout_seconds: float):
        self._base_url = base_url.rstrip("/")
        self._timeout_seconds = timeout_seconds

    async def create(self, *, model: str, messages: List[Dict[str, str]], temperature: float, max_tokens: int):
        def _post():
            headers = {"Content-Type": "application/json"}

            attempts = []
            if self._base_url.endswith("/v1"):
                attempts.append((f"{self._base_url}/chat/completions", "openai"))
            else:
                attempts.append((f"{self._base_url}/api/chat", "ollama"))
                attempts.append((f"{self._base_url}/v1/chat/completions", "openai"))

            last_exc = None
            for url, mode in attempts:
                if mode == "ollama":
                    payload = json.dumps(
                        {
                            "model": model,
                            "messages": messages,
                            "stream": False,
                            "options": {
                                "temperature": temperature,
                                "num_predict": max_tokens,
                            },
                        }
                    ).encode("utf-8")
                else:
                    payload = json.dumps(
                        {
                            "model": model,
                            "messages": messages,
                            "temperature": temperature,
                            "max_tokens": max_tokens,
                        }
                    ).encode("utf-8")

                req = urllib_request.Request(
                    url,
                    data=payload,
                    headers=headers,
                    method="POST",
                )
                try:
                    with urllib_request.urlopen(req, timeout=self._timeout_seconds) as resp:
                        raw = resp.read().decode("utf-8")
                    data = json.loads(raw)
                    if mode == "openai":
                        content = data["choices"][0]["message"]["content"]
                    else:
                        content = data["message"]["content"]
                    return SimpleNamespace(
                        choices=[SimpleNamespace(message=SimpleNamespace(content=content))]
                    )
                except Exception as exc:
                    last_exc = exc
                    continue

            raise RuntimeError(f"Failed to contact Ollama at {self._base_url}: {last_exc}") from last_exc

        return await asyncio.to_thread(_post)


class _OllamaClient:
    def __init__(self, base_url: str, timeout_seconds: float):
        self.chat = SimpleNamespace(completions=_OllamaChatCompletions(base_url, timeout_seconds))


OLLAMA_BASE_URL = os.getenv("OLLAMA_BASE_URL", "http://localhost:11434/v1")
OLLAMA_TIMEOUT_SECONDS = float(os.getenv("OLLAMA_TIMEOUT_SECONDS", "60"))
llm_client = _OllamaClient(base_url=OLLAMA_BASE_URL, timeout_seconds=OLLAMA_TIMEOUT_SECONDS)

# ---------------------------------------------------------------
# OCR — lazy import
# ---------------------------------------------------------------
try:
    from hybrid_ocr_monopoly import process_image, get_easyocr_reader
    OCR_AVAILABLE = True
except ImportError:
    OCR_AVAILABLE = False
    logger.warning("OCR module not available — /ocr endpoint will be disabled.")

# ---------------------------------------------------------------
# Quiz — lazy import so API startup does not fail if quiz deps are missing
# ---------------------------------------------------------------
quiz_sessions: Dict[str, Any] = {}
QUIZ_AVAILABLE = False
_question_bank = None
_bandit = None
_forum = None

try:
    from adaptive_quiz import QuestionBank, ThompsonBandit, PracticeForum

    _question_bank = QuestionBank()
    _bandit = ThompsonBandit(_question_bank.all_topics)
    _bandit.load_profile("bayes_model.json")   # loads saved progress if it exists
    _forum = PracticeForum()
    QUIZ_AVAILABLE = True
except ImportError as exc:
    logger.warning(f"Quiz module unavailable — quiz endpoints will be disabled: {exc}")

# ---------------------------------------------------------------
# System prompt
# ---------------------------------------------------------------
CONTEST_BANK_AVAILABLE = False
_contest_problem_bank: List[Dict[str, Any]] = []

try:
    from openpyxl import load_workbook

    contest_bank_path = os.getenv("OLYMPIAD_XLSX_PATH", "olympiad.xlsx")
    workbook = load_workbook(contest_bank_path, data_only=True)
    worksheet = workbook["Olympiad Questions"]

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue

        try:
            difficulty_1_to_10 = max(1, min(10, int((row[7] if len(row) > 7 else 5) or 5)))
        except Exception:
            difficulty_1_to_10 = 5

        answer_value = row[6]
        if answer_value is None:
            continue

        answer_match = re.search(r"-?\d+", str(answer_value))
        if not answer_match:
            continue

        normalized_answer = int(answer_match.group(0))
        if normalized_answer < 0 or normalized_answer > 99:
            normalized_answer %= 100

        _contest_problem_bank.append(
            {
                "questionId": str(row[0]),
                "competition": str(row[1] or "Olympiad"),
                "year": str(row[2] or ""),
                "level": str(row[3] or ""),
                "topic": str(row[4] or "Olympiad"),
                "statement": str(row[5] or "").strip(),
                "correctAnswer": normalized_answer,
                "difficultyInput": difficulty_1_to_10,
                "difficultyRating": int(800 + (difficulty_1_to_10 - 1) * 177.78),
                "source": str(row[8] or row[1] or "Olympiad Workbook"),
            }
        )

    CONTEST_BANK_AVAILABLE = len(_contest_problem_bank) > 0
except Exception as exc:
    logger.warning(f"Contest problem bank unavailable â€” contest endpoints will be limited: {exc}")

SYSTEM_PROMPT = """You are MathMend, an expert math tutor and solver.

When given a math problem:
1. Identify all variables and what they represent.
2. Write out the equations that model the problem.
3. Solve the equations step-by-step.
4. State the final answer clearly.

Format your response EXACTLY like this example — no markdown, no bullet points:

A photo is 8 inches by 6 inches. If you enlarge it by a scale factor of 1.5, what are the new dimensions?
The original height is 6 inches. The original width is 8 inches. After enlargement, the height is h inches. After enlargement, the width is w inches.  Scale factor relates old and new heights: h = 6 * 1.5 Scale factor relates old and new widths: w = 8 * 1.5  Output: h = 6 * 1.5 w = 8 * 1.5  Solution: h = 9.0, w = 12.0

Key rules:
- Always end with "Solution: variable = value, variable = value"
- Use plain text only, no markdown headers or bullet points
- Keep the explanation concise but complete
- If the problem is simple arithmetic, still show the equation then the answer
"""

# ---------------------------------------------------------------
# Lifespan
# ---------------------------------------------------------------
@asynccontextmanager
async def lifespan(app: FastAPI):
    logger.info("🚀 MathMend API starting up...")
    if OCR_AVAILABLE:
        try:
            await asyncio.to_thread(get_easyocr_reader)
            logger.info("✅ OCR engine ready.")
        except Exception as e:
            logger.warning(f"⚠️ OCR warmup failed: {e}")
    logger.info("✅ MathMend API ready!")
    yield
    logger.info("🛑 MathMend API shutting down.")

# ---------------------------------------------------------------
# App
# ---------------------------------------------------------------
app = FastAPI(
    title="MathMend API",
    description="Math solver + Adaptive Quiz API",
    version="4.0.0",
    lifespan=lifespan,
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000", "http://127.0.0.1:3000", "*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ---------------------------------------------------------------
# Pydantic models — Chat
# ---------------------------------------------------------------
class ChatRequest(BaseModel):
    question: str = Field(..., min_length=1)

class ChatResponse(BaseModel):
    answer: str
    solution: Dict[str, Any] = {}
    reasoning: Optional[str] = None
    equations: List[str] = []
    latency_ms: float
    result_type: str = "llm"

class OCRResponse(BaseModel):
    extracted_text: str
    latency_ms: float


def _require_quiz():
    if not QUIZ_AVAILABLE:
        raise HTTPException(status_code=503, detail="Quiz module dependencies are not installed on this server.")

# ---------------------------------------------------------------
# Pydantic models — Quiz
# ---------------------------------------------------------------
class QuizStartRequest(BaseModel):
    batch_size: int = Field(default=5, ge=1, le=20)

class QuizAnswerRequest(BaseModel):
    session_id: str
    question_index: int
    answer: str
    topic: str
    ground_truth: float

class QuizQuestion(BaseModel):
    topic: str
    question: str
    solution: str
    ground_truth: float

class QuizStartResponse(BaseModel):
    session_id: str
    questions: List[QuizQuestion]

class QuizAnswerResponse(BaseModel):
    correct: bool
    solution: str

class TopicStats(BaseModel):
    mean: float
    certainty: float

class QuizStatsResponse(BaseModel):
    topics: Dict[str, TopicStats]


class ContestProblem(BaseModel):
    questionId: str
    competition: str
    year: str
    level: str
    topic: str
    statement: str
    correctAnswer: int = Field(..., ge=0, le=99)
    difficultyInput: int = Field(..., ge=1, le=10)
    difficultyRating: int = Field(..., ge=800, le=2600)
    source: str


class ContestProblemBankResponse(BaseModel):
    total: int
    problems: List[ContestProblem]


class ContestQuestion(BaseModel):
    questionId: str
    statement: str
    correctAnswer: int = Field(..., ge=0, le=99)
    difficultyRating: int = Field(..., ge=800, le=2600)
    topic: str = "Olympiad"


class ContestSnapshot(BaseModel):
    contestId: str
    title: str
    visibility: str
    startTime: str
    endTime: str
    duration: int = Field(..., ge=60, le=21600)
    questions: List[ContestQuestion]


class ContestAnswerPayload(BaseModel):
    qId: str
    answer: Optional[int] = Field(default=None, ge=0, le=99)
    timestamp: str


class ContestSubmissionPayload(BaseModel):
    userId: str
    displayName: Optional[str] = None
    startTime: str
    endTime: str
    answers: List[ContestAnswerPayload]
    currentRating: float = 0
    maxRating: float = 0
    experience: int = 0
    streak: int = 0
    volatility: float = 1.0


class ContestEvaluationRequest(BaseModel):
    contest: ContestSnapshot
    submissions: List[ContestSubmissionPayload]


class ContestEvaluationResponse(BaseModel):
    contestId: str
    leaderboard: List[Dict[str, Any]]
    evaluations: List[Dict[str, Any]]
    audit: Dict[str, Any]

# ---------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------
def _parse_solution_from_text(text: str) -> Dict[str, Any]:
    solution = {}
    match = re.search(r"Solution:\s*(.+)", text, re.IGNORECASE)
    if not match:
        return solution
    pairs_str = match.group(1)
    for pair in re.finditer(r"(\w+)\s*=\s*([\d.]+)", pairs_str):
        key = pair.group(1)
        try:
            solution[key] = float(pair.group(2))
        except ValueError:
            solution[key] = pair.group(2)
    return solution

def _parse_equations_from_text(text: str) -> List[str]:
    equations = []
    for line in text.split("\n"):
        line = line.strip()
        if re.search(r"[\w\s]+[+\-*/^][\w\s]+=\s*[\d\w]", line):
            if len(line) < 120:
                equations.append(line)
    return equations[:6]


def _extract_math_expression(question: str) -> str:
    cleaned = question.strip()
    cleaned = re.sub(
        r"^(simplify|solve|evaluate|calculate|compute|find|what is|what's)\b[:\-\s]*",
        "",
        cleaned,
        flags=re.IGNORECASE,
    )
    cleaned = cleaned.split("?")[0].split(".")[0]
    expr = re.sub(r"[^0-9\.\+\-\*\/\^\(\)\s]", "", cleaned)
    expr = re.sub(r"\s+", "", expr).replace("^", "**")
    return expr


def _safe_eval_math_expression(expr: str):
    def _eval(node):
        if isinstance(node, ast.Expression):
            return _eval(node.body)
        if isinstance(node, ast.Constant):
            value = node.value
            if isinstance(value, (int, float)):
                return Fraction(str(value))
            raise ValueError("Unsupported constant")
        if isinstance(node, ast.BinOp):
            left = _eval(node.left)
            right = _eval(node.right)
            if isinstance(node.op, ast.Add):
                return left + right
            if isinstance(node.op, ast.Sub):
                return left - right
            if isinstance(node.op, ast.Mult):
                return left * right
            if isinstance(node.op, ast.Div):
                return left / right
            if isinstance(node.op, ast.Pow):
                if right.denominator != 1:
                    raise ValueError("Non-integer exponent")
                return left ** int(right)
            raise ValueError("Unsupported operator")
        if isinstance(node, ast.UnaryOp):
            operand = _eval(node.operand)
            if isinstance(node.op, ast.UAdd):
                return operand
            if isinstance(node.op, ast.USub):
                return -operand
            raise ValueError("Unsupported unary operator")
        raise ValueError("Unsupported expression")

    parsed = ast.parse(expr, mode="eval")
    return _eval(parsed)


def _format_number(value):
    if isinstance(value, Fraction):
        if value.denominator == 1:
            return str(value.numerator)
        return f"{value.numerator}/{value.denominator} ({float(value):.6g})"
    if isinstance(value, float):
        return f"{value:.6g}"
    return str(value)


def _fallback_math_answer(question: str, latency_ms: float) -> ChatResponse:
    expr = _extract_math_expression(question)
    if expr and re.search(r"\d", expr):
        try:
            value = _safe_eval_math_expression(expr)
            if isinstance(value, Fraction) and value.denominator != 1:
                solution_value = float(value)
            else:
                solution_value = float(value)
            answer_text = (
                f"The expression is {expr}. "
                f"Using order of operations, the value is {_format_number(value)}. "
                f"Solution: value = {_format_number(value)}"
            )
            return ChatResponse(
                answer=answer_text,
                solution={"value": solution_value},
                reasoning=answer_text,
                equations=[f"{expr} = {_format_number(value)}"],
                latency_ms=latency_ms,
                result_type="fallback",
            )
        except Exception as exc:
            logger.warning(f"Fallback math parser could not solve expression '{expr}': {exc}")

    answer_text = (
        "I could not reach the Ollama model on this machine, so I cannot produce a full model answer right now. "
        "Please try a simpler arithmetic expression or install a smaller Ollama model. "
        "Solution: value = unavailable"
    )
    return ChatResponse(
        answer=answer_text,
        solution={},
        reasoning=answer_text,
        equations=[],
        latency_ms=latency_ms,
        result_type="fallback",
    )


def _require_contest_bank():
    if not CONTEST_BANK_AVAILABLE:
        raise HTTPException(status_code=503, detail="Contest problem bank is not available on this server.")


def _parse_iso_timestamp(value: str) -> float:
    normalized = (value or "").strip().replace("Z", "+00:00")
    return __import__("datetime").datetime.fromisoformat(normalized).timestamp()


def _normalize_contest_answer(value: Optional[int]) -> Optional[int]:
    if value is None:
        return None
    value = int(value)
    if value < 0 or value > 99:
        value %= 100
    return value


def _clamp(value: float, lower: float, upper: float) -> float:
    return max(lower, min(upper, value))


def _detect_pairwise_patterns(evaluations: List[Dict[str, Any]]) -> Dict[str, List[str]]:
    flags: Dict[str, List[str]] = {entry["userId"]: [] for entry in evaluations}

    for index, left in enumerate(evaluations):
        left_answers = [item.get("answer") for item in left["answers"]]
        left_time_marks = [item.get("offsetSeconds") for item in left["answers"]]

        for right in evaluations[index + 1 :]:
            right_answers = [item.get("answer") for item in right["answers"]]
            right_time_marks = [item.get("offsetSeconds") for item in right["answers"]]

            if left_answers and left_answers == right_answers:
                average_offset_gap = (
                    mean(
                        [abs((left_time_marks[pos] or 0) - (right_time_marks[pos] or 0)) for pos in range(len(left_time_marks))]
                    )
                    if left_time_marks and right_time_marks and len(left_time_marks) == len(right_time_marks)
                    else 999
                )
                if average_offset_gap <= 3:
                    flags[left["userId"]].append(f"Matched answer vector with {right['userId']} on near-identical timing.")
                    flags[right["userId"]].append(f"Matched answer vector with {left['userId']} on near-identical timing.")

    return flags


def _build_contest_evaluation(request: ContestEvaluationRequest) -> Dict[str, Any]:
    contest = request.contest
    start_ts = _parse_iso_timestamp(contest.startTime)
    end_ts = _parse_iso_timestamp(contest.endTime)
    questions_by_id = {question.questionId: question for question in contest.questions}
    total_question_weight = sum(question.difficultyRating for question in contest.questions) or 1
    average_participant_rating = (
        mean([submission.currentRating for submission in request.submissions]) if request.submissions else 0
    )

    evaluations: List[Dict[str, Any]] = []

    for submission in request.submissions:
        submission_start_ts = _parse_iso_timestamp(submission.startTime)
        submission_end_ts = _parse_iso_timestamp(submission.endTime)
        ordered_answers = sorted(submission.answers, key=lambda item: _parse_iso_timestamp(item.timestamp))
        answer_map = {answer.qId: answer for answer in ordered_answers}
        normalized_total_time = max(1, int(submission_end_ts - submission_start_ts))
        per_question_seconds = normalized_total_time / max(len(contest.questions), 1)
        difficulty_solved = 0
        total_correct = 0
        answer_rows = []
        unique_answers = set()

        for question in contest.questions:
            raw_answer = answer_map.get(question.questionId)
            normalized_answer = _normalize_contest_answer(raw_answer.answer if raw_answer else None)
            is_correct = normalized_answer == question.correctAnswer if normalized_answer is not None else False
            timestamp = raw_answer.timestamp if raw_answer else submission.endTime
            answer_ts = _parse_iso_timestamp(timestamp)
            offset_seconds = max(0, int(answer_ts - submission_start_ts))
            total_correct += 1 if is_correct else 0
            difficulty_solved += question.difficultyRating if is_correct else 0
            if normalized_answer is not None:
                unique_answers.add(normalized_answer)

            answer_rows.append(
                {
                    "qId": question.questionId,
                    "answer": normalized_answer,
                    "timestamp": timestamp,
                    "offsetSeconds": offset_seconds,
                    "isCorrect": is_correct,
                    "difficultyRating": question.difficultyRating,
                    "topic": question.topic,
                }
            )

        weighted_accuracy = difficulty_solved / total_question_weight
        actual_score = total_correct / max(len(contest.questions), 1)
        expected_score = 1 / (1 + 10 ** ((average_participant_rating - submission.currentRating) / 400))
        dynamic_k = 56 if submission.experience < 5 else 40 if submission.experience < 15 else 28
        volatility_factor = _clamp(1 + submission.volatility * 0.25, 0.9, 1.45)
        normalized_time = _clamp(normalized_total_time / max(contest.duration, 1), 0.05, 1.5)
        speed_factor = math.exp(-0.65 * normalized_time)
        time_penalty = 42 * math.log1p(normalized_total_time)
        streak_bonus = 6 * max(submission.streak, 0)
        difficulty_boost = sum(
            item["difficultyRating"] * (0.018 if item["isCorrect"] else 0.0)
            for item in answer_rows
        )
        performance_score = difficulty_solved - time_penalty + streak_bonus
        rating_change = (
            (dynamic_k * (actual_score - expected_score) * volatility_factor * max(speed_factor, 0.42))
            + difficulty_boost
            + min(streak_bonus, 24)
            + (weighted_accuracy - 0.5) * 34
        )

        fairness_flags = []
        valid = True

        if submission_start_ts < start_ts or submission_start_ts > end_ts:
            fairness_flags.append("Contest was not started inside the allowed contest window.")
            valid = False
        if submission_end_ts > end_ts + 5:
            fairness_flags.append("Submission closed after the contest end time.")
            valid = False
        if normalized_total_time > contest.duration + 5:
            fairness_flags.append("Submission duration exceeded the contest limit.")
            valid = False
        if per_question_seconds < 5:
            fairness_flags.append("Submission speed fell below the minimum human threshold.")
            valid = False
        if weighted_accuracy >= 0.9 and submission.currentRating < average_participant_rating - 250:
            fairness_flags.append("Accuracy spike is inconsistent with the participant rating baseline.")
        if len(unique_answers) <= max(1, len(contest.questions) // 4) and total_correct <= max(1, len(contest.questions) // 4):
            fairness_flags.append("Answer pattern resembles low-entropy guessing.")

        if not valid:
            rating_change = 0

        new_rating = max(800, round(submission.currentRating + rating_change))
        evaluations.append(
            {
                "userId": submission.userId,
                "displayName": submission.displayName or submission.userId,
                "answers": answer_rows,
                "startTime": submission.startTime,
                "endTime": submission.endTime,
                "totalCorrect": total_correct,
                "difficultySolved": difficulty_solved,
                "totalTime": normalized_total_time,
                "penalties": round(time_penalty, 2),
                "performanceScore": round(performance_score, 2),
                "actualScore": round(actual_score, 4),
                "expectedScore": round(expected_score, 4),
                "speedFactor": round(speed_factor, 4),
                "ratingChange": round(rating_change),
                "newRating": new_rating,
                "volatility": round(_clamp(submission.volatility * 0.97 + (0.12 if valid else 0.2), 0.5, 1.8), 3),
                "streak": submission.streak,
                "experience": submission.experience,
                "isValid": valid,
                "fairnessFlags": fairness_flags,
            }
        )

    pairwise_flags = _detect_pairwise_patterns(evaluations)
    for evaluation in evaluations:
        if pairwise_flags[evaluation["userId"]]:
            evaluation["fairnessFlags"].extend(pairwise_flags[evaluation["userId"]])
            evaluation["isValid"] = False
            evaluation["ratingChange"] = 0
            evaluation["newRating"] = max(800, round(request.submissions[[s.userId for s in request.submissions].index(evaluation["userId"])].currentRating))

    valid_entries = [entry for entry in evaluations if entry["isValid"]]
    valid_entries.sort(
        key=lambda item: (
            -item["totalCorrect"],
            item["totalTime"],
            -item["difficultySolved"],
            item["userId"],
        )
    )

    leaderboard = []
    for rank, entry in enumerate(valid_entries, start=1):
        leaderboard.append(
            {
                "userId": entry["userId"],
                "displayName": entry["displayName"],
                "score": entry["totalCorrect"],
                "rank": rank,
                "ratingChange": entry["ratingChange"],
                "newRating": entry["newRating"],
                "totalTime": entry["totalTime"],
                "difficultySolved": entry["difficultySolved"],
                "performanceScore": entry["performanceScore"],
            }
        )

    for entry in evaluations:
        if not entry["isValid"]:
            leaderboard.append(
                {
                    "userId": entry["userId"],
                    "displayName": entry["displayName"],
                    "score": entry["totalCorrect"],
                    "rank": None,
                    "ratingChange": 0,
                    "newRating": entry["newRating"],
                    "totalTime": entry["totalTime"],
                    "difficultySolved": entry["difficultySolved"],
                    "performanceScore": entry["performanceScore"],
                    "disqualified": True,
                }
            )

    audit_payload = {
        "contestId": contest.contestId,
        "generatedAt": time.time(),
        "participantCount": len(request.submissions),
        "leaderboard": leaderboard,
        "fairnessSummary": [
            {
                "userId": entry["userId"],
                "isValid": entry["isValid"],
                "flags": entry["fairnessFlags"],
            }
            for entry in evaluations
        ],
    }
    audit_hash = hashlib.sha256(json.dumps(audit_payload, sort_keys=True).encode("utf-8")).hexdigest()

    return {
        "contestId": contest.contestId,
        "leaderboard": leaderboard,
        "evaluations": evaluations,
        "audit": {
            "hash": audit_hash,
            "generatedAt": __import__("datetime").datetime.utcnow().isoformat() + "Z",
            "participantCount": len(request.submissions),
            "validParticipants": len(valid_entries),
        },
    }

# ---------------------------------------------------------------
# Routes — General
# ---------------------------------------------------------------
@app.get("/health")
async def health_check():
    return {
        "status": "healthy",
        "model": MODEL_NAME,
        "ocr_available": OCR_AVAILABLE,
        "quiz_available": QUIZ_AVAILABLE,
        "contest_bank_available": CONTEST_BANK_AVAILABLE,
    }


@app.get("/contest/problems", response_model=ContestProblemBankResponse)
async def contest_problem_bank(
    limit: int = 40,
    search: Optional[str] = None,
    topic: Optional[str] = None,
    level: Optional[str] = None,
):
    _require_contest_bank()
    filtered = _contest_problem_bank

    if search:
        search_lower = search.lower()
        filtered = [
            item
            for item in filtered
            if search_lower in item["statement"].lower()
            or search_lower in item["topic"].lower()
            or search_lower in item["competition"].lower()
        ]
    if topic:
        filtered = [item for item in filtered if item["topic"].lower() == topic.lower()]
    if level:
        filtered = [item for item in filtered if item["level"].lower() == level.lower()]

    bounded_limit = max(1, min(limit, 120))
    return ContestProblemBankResponse(total=len(filtered), problems=filtered[:bounded_limit])


@app.get("/contest/server-time")
async def contest_server_time():
    return {
        "serverTime": __import__("datetime").datetime.utcnow().isoformat() + "Z",
        "epochMs": int(time.time() * 1000),
    }


@app.post("/contest/evaluate", response_model=ContestEvaluationResponse)
async def evaluate_contest(request: ContestEvaluationRequest):
    if not request.contest.questions:
        raise HTTPException(status_code=400, detail="Contest requires at least one question.")
    if not request.submissions:
        raise HTTPException(status_code=400, detail="Contest evaluation requires at least one submission.")

    try:
        result = _build_contest_evaluation(request)
        return ContestEvaluationResponse(**result)
    except HTTPException:
        raise
    except Exception as exc:
        logger.error(f"Contest evaluation failed: {exc}", exc_info=True)
        raise HTTPException(status_code=500, detail="Contest evaluation failed.")

# ---------------------------------------------------------------
# Routes — Chat
# ---------------------------------------------------------------
@app.post("/chat", response_model=ChatResponse)
async def chat_endpoint(req: ChatRequest):
    start_time = time.time()
    question = req.question.strip()
    logger.info(f"Chat request: {question[:80]}...")

    try:
        response = await llm_client.chat.completions.create(
            model=MODEL_NAME,
            messages=[
                {"role": "system", "content": SYSTEM_PROMPT},
                {"role": "user",   "content": question},
            ],
            temperature=0.2,
            max_tokens=800,
        )
        answer_text = response.choices[0].message.content.strip()
        solution    = _parse_solution_from_text(answer_text)
        equations   = _parse_equations_from_text(answer_text)
        latency     = (time.time() - start_time) * 1000
        logger.info(f"Answered in {latency:.1f}ms | solution keys: {list(solution.keys())}")
        return ChatResponse(
            answer=answer_text,
            solution=solution,
            reasoning=answer_text,
            equations=equations,
            latency_ms=latency,
            result_type="llm",
        )
    except Exception as e:
        logger.warning(f"LLM call failed, using fallback: {e}", exc_info=True)
        fallback = _fallback_math_answer(question, (time.time() - start_time) * 1000)
        logger.info(f"Fallback response used | result_type={fallback.result_type}")
        return fallback

# ---------------------------------------------------------------
# Routes — OCR
# ---------------------------------------------------------------
@app.post("/ocr", response_model=OCRResponse)
async def ocr_endpoint(image: UploadFile = File(...)):
    if not OCR_AVAILABLE:
        raise HTTPException(status_code=503, detail="OCR module not available on this server.")
    if not image:
        raise HTTPException(status_code=400, detail="No image provided")

    start_time = time.time()
    suffix = os.path.splitext(image.filename or "upload.png")[1] or ".png"
    tmp_path = ""

    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            content = await image.read()
            tmp.write(content)
            tmp_path = tmp.name
        extracted_text = await asyncio.to_thread(process_image, tmp_path)
        latency = (time.time() - start_time) * 1000
        logger.info(f"OCR completed in {latency:.1f}ms")
        return OCRResponse(extracted_text=extracted_text, latency_ms=latency)
    except Exception as e:
        logger.error(f"OCR error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"OCR error: {str(e)}")
    finally:
        if tmp_path and os.path.exists(tmp_path):
            try:
                os.remove(tmp_path)
            except Exception:
                pass

# ---------------------------------------------------------------
# Routes — Quiz
# ---------------------------------------------------------------
@app.post("/quiz/start", response_model=QuizStartResponse)
async def quiz_start(req: QuizStartRequest):
    _require_quiz()
    topics = _bandit.select_topics(k=req.batch_size)
    questions_out = []
    raw_questions = []

    for topic in topics:
        q_data = _question_bank.get_question(topic)
        if not q_data:
            continue
        questions_out.append(QuizQuestion(
            topic=topic,
            question=q_data["q"],
            solution=q_data["solution_text"],
            ground_truth=float(q_data["ground_truth_numeric"]),
        ))
        raw_questions.append(q_data)

    session_id = str(uuid.uuid4())
    quiz_sessions[session_id] = {
        "topics": topics,
        "raw_questions": raw_questions,
    }
    logger.info(f"Quiz session started: {session_id} | {len(questions_out)} questions")
    return QuizStartResponse(session_id=session_id, questions=questions_out)


@app.post("/quiz/answer", response_model=QuizAnswerResponse)
async def quiz_answer(req: QuizAnswerRequest):
    _require_quiz()
    if req.session_id not in quiz_sessions:
        raise HTTPException(status_code=404, detail="Session not found")

    session = quiz_sessions[req.session_id]
    is_correct = await asyncio.to_thread(
        _forum.grade_submission,
        req.answer,
        req.ground_truth,
    )
    _bandit.update(req.topic, is_correct)

    idx = req.question_index
    raw_questions = session.get("raw_questions", [])
    solution_text = (
        raw_questions[idx]["solution_text"]
        if idx < len(raw_questions)
        else "See solution above."
    )
    logger.info(f"Answer graded | correct={is_correct} | topic={req.topic}")
    return QuizAnswerResponse(correct=is_correct, solution=solution_text)


@app.get("/quiz/stats", response_model=QuizStatsResponse)
async def quiz_stats():
    _require_quiz()
    result = {}
    for topic in _question_bank.all_topics:
        mean, certainty = _bandit.get_stats(topic)
        result[topic] = TopicStats(mean=mean, certainty=certainty)
    return QuizStatsResponse(topics=result)


if __name__ == "__main__":
    uvicorn.run(app, host="localhost", port=8000, http="h11", ws="websockets")
